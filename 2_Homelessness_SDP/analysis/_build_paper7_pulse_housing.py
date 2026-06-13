"""Build tidy Household Pulse housing precarity panel (paper7).
MECHANICAL: reads only downloaded xlsx; no hand-coded values.

housing3b: 'Likelihood of Having to Leave ... Due to Eviction' (universe=renters NOT current on rent)
  Total row cols: 1=Total, 2=Very likely, 3=Somewhat likely, 4=Not very likely, 5=Not likely at all, 6=Did not report
  eviction_risk_share = (Very likely + Somewhat likely) / Total
housing1b: 'Last Month's Payment Status for Renter-Occupied HU' (universe=all renters 18+)
  Total row cols: 1=Total, 2=caught up Yes, 3=caught up No, 4=caught up Did not report, 5=occupied w/o rent, 6=DNR tenure
  behind_on_rent_share = (caught up No) / Total
Anchor = first data row where col0.strip()=='Total'. '-' treated as missing.
"""
import openpyxl, json, glob, os, re

PULSE_DIR = r"D:/IDP/data/paper7/pulse"
OUT_CSV = r"D:/IDP/analysis/paper7_pulse_housing_precarity.csv"
OUT_JSON = r"D:/IDP/analysis/paper7_pulse_coverage.json"

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s in ('-', '', 'N', 'NA', '*'): return None
    s = s.replace(',', '')
    try: return float(s)
    except ValueError: return None

def find_total_row(ws):
    """Return the values tuple of the first row whose col0 stripped == 'Total'."""
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        c0 = row[0]
        if c0 is not None and str(c0).strip() == 'Total':
            return row
    return None

def geo_type(sheet):
    if sheet == 'US': return 'US'
    if sheet.endswith('_Metro_Area'): return 'metro'
    return 'state'

def clean_geo(sheet):
    if sheet.endswith('_Metro_Area'):
        name = sheet[:-len('_Metro_Area')].replace('.', ' ').replace('_', ' ')
        return name + ' Metro Area'
    return sheet

def parse_file(path, table):
    """Return dict {sheet: value} of the metric for the given table ('3b' or '1b')."""
    out = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        tr = find_total_row(ws)
        if tr is None:
            out[sheet] = (None, 'no_total_row')
            continue
        try:
            total = num(tr[1])
            if table == '3b':
                vl, sl = num(tr[2]), num(tr[3])
                if total in (None, 0) or vl is None or sl is None:
                    out[sheet] = (None, 'missing_or_zero'); continue
                out[sheet] = ((vl + sl) / total, None)
            else:  # 1b
                behind = num(tr[3])
                if total in (None, 0) or behind is None:
                    out[sheet] = (None, 'missing_or_zero'); continue
                out[sheet] = (behind / total, None)
        except (IndexError, TypeError) as e:
            out[sheet] = (None, f'err:{e}')
    wb.close()
    return out

# Enumerate periods present
periods = {}  # key -> dict(year, period_label, cycle_or_week, f3b, f1b)

# 2024 cycles
for f in glob.glob(os.path.join(PULSE_DIR, 'housing3b_cycle*.xlsx')):
    m = re.search(r'cycle(\d{2})\.xlsx', f)
    if not m: continue
    nn = m.group(1)
    key = f"2024-cycle{nn}"
    periods.setdefault(key, {})
    periods[key].update(year=2024, period_label=f"2024 Cycle {int(nn)}",
                         cycle_or_week=f"cycle{nn}", f3b=f)
for f in glob.glob(os.path.join(PULSE_DIR, 'housing1b_cycle*.xlsx')):
    m = re.search(r'cycle(\d{2})\.xlsx', f)
    if not m: continue
    nn = m.group(1)
    key = f"2024-cycle{nn}"
    periods.setdefault(key, {})
    periods[key].setdefault('year', 2024)
    periods[key].setdefault('period_label', f"2024 Cycle {int(nn)}")
    periods[key].setdefault('cycle_or_week', f"cycle{nn}")
    periods[key]['f1b'] = f

# 2022 / 2023 weeks
for f in glob.glob(os.path.join(PULSE_DIR, 'housing3b_20*_week*.xlsx')):
    m = re.search(r'housing3b_(\d{4})_week(\d+)\.xlsx', f)
    if not m: continue
    yr, wk = m.group(1), m.group(2)
    key = f"{yr}-wk{wk}"
    periods.setdefault(key, {})
    periods[key].update(year=int(yr), period_label=f"{yr} Week {wk}",
                        cycle_or_week=f"wk{wk}", f3b=f)
for f in glob.glob(os.path.join(PULSE_DIR, 'housing1b_20*_week*.xlsx')):
    m = re.search(r'housing1b_(\d{4})_week(\d+)\.xlsx', f)
    if not m: continue
    yr, wk = m.group(1), m.group(2)
    key = f"{yr}-wk{wk}"
    periods.setdefault(key, {})
    periods[key].setdefault('year', int(yr))
    periods[key].setdefault('period_label', f"{yr} Week {wk}")
    periods[key].setdefault('cycle_or_week', f"wk{wk}")
    periods[key]['f1b'] = f

# sort: year then cycle/week number
def sortkey(k):
    p = periods[k]
    m = re.search(r'(\d+)', p['cycle_or_week'])
    return (p['year'], int(m.group(1)) if m else 0)

rows = []
coverage = {'periods': [], 'caveats': [], 'parse_errors': []}

for key in sorted(periods, key=sortkey):
    p = periods[key]
    f3b = p.get('f3b'); f1b = p.get('f1b')
    e_map = parse_file(f3b, '3b') if f3b else {}
    b_map = parse_file(f1b, '1b') if f1b else {}
    sheets = sorted(set(e_map) | set(b_map))
    ngeo = 0
    for sheet in sheets:
        ev = e_map.get(sheet, (None, 'no_3b_file'))
        bh = b_map.get(sheet, (None, 'no_1b_file'))
        ev_val, ev_err = ev if isinstance(ev, tuple) else (ev, None)
        bh_val, bh_err = bh if isinstance(bh, tuple) else (bh, None)
        if ev_err: coverage['parse_errors'].append(f"{key}|3b|{sheet}|{ev_err}")
        if bh_err: coverage['parse_errors'].append(f"{key}|1b|{sheet}|{bh_err}")
        rows.append({
            'cycle_or_week': p['cycle_or_week'],
            'period_label': p['period_label'],
            'year': p['year'],
            'geography': clean_geo(sheet),
            'geo_type': geo_type(sheet),
            'eviction_risk_share': round(ev_val, 6) if ev_val is not None else '',
            'behind_on_rent_share': round(bh_val, 6) if bh_val is not None else '',
        })
        ngeo += 1
    coverage['periods'].append({
        'key': key, 'year': p['year'], 'period_label': p['period_label'],
        'cycle_or_week': p['cycle_or_week'],
        'has_housing3b': bool(f3b), 'has_housing1b': bool(f1b),
        'n_geographies': ngeo,
    })

# write CSV
import csv
fields = ['cycle_or_week','period_label','year','geography','geo_type',
          'eviction_risk_share','behind_on_rent_share']
with open(OUT_CSV, 'w', newline='', encoding='utf-8') as fh:
    w = csv.DictWriter(fh, fieldnames=fields)
    w.writeheader()
    w.writerows(rows)

coverage['n_rows'] = len(rows)
coverage['n_periods'] = len(periods)
coverage['caveats'] = [
    "eviction_risk_share = (Very likely + Somewhat likely)/Total from housing3b 'Total' row; "
    "housing3b universe = renters 18+ NOT current on rental payments (i.e. share of behind renters who expect eviction).",
    "behind_on_rent_share = (caught up 'No')/Total from housing1b 'Total' row; "
    "housing1b universe = all renter-occupied HU 18+.",
    "Anchor row = first row where col0=='Total'. Column positions: 3b cols 2,3=Very/Somewhat likely, col1=Total; "
    "1b col3='No' caught up, col1=Total. Verified header rows 4-5 stable across 2022/2023/2024 files.",
    "Cross-check: housing1b 'No' count == housing3b 'Total' count (same population), confirmed for US 2022 wk42 & 2024 cycle09.",
    "'-' suppressed cells treated as missing (blank in output).",
    "2022=weeks 42-51, 2023=weeks 52-63, 2024=cycles 01-09 (cycle naming replaced week naming in 2024).",
]
with open(OUT_JSON, 'w', encoding='utf-8') as fh:
    json.dump(coverage, fh, indent=2)

print("Rows:", len(rows), "Periods:", len(periods))
print("Parse errors:", len(coverage['parse_errors']))
